import openpyxl
import requests
import os
import zipfile
import shutil
import re
from collections import defaultdict
import logging
from datetime import datetime
import concurrent.futures
import argparse

# At the beginning of your script, add:
logger = logging.getLogger(__name__)

def download_file(url, filename):
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36',
    }
    response = requests.get(url, headers=headers)
    if response.status_code == 200:
        with open(filename, 'wb') as file:
            file.write(response.content)
        return True
    return False

def extract_download_link(html_url):
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36',
    }
    
    response = requests.get(html_url, headers=headers)
    if response.status_code == 200:
        # Use regex to find the download link in the JavaScript
        match = re.search(r"window\.location\.href='(https?://.*?\.zip)'", response.text)
        if match:
            return match.group(1)
        else:
            print(f"No download link found in the HTML from: {html_url}")
    else:
        print(f"Failed to retrieve HTML. Status code: {response.status_code}")
    return None

def extract_zip(zip_file, extract_to):
    with zipfile.ZipFile(zip_file, 'r') as zip_ref:
        zip_ref.extractall(extract_to)

def setup_logging():
    # Create a logs directory if it doesn't exist
    os.makedirs('logs', exist_ok=True)

    # Set up logging
    log_filename = f"logs/download_log_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s',
        handlers=[
            logging.FileHandler(log_filename),
            logging.StreamHandler()
        ]
    )
    return logging.getLogger(__name__)

def process_tdoc(row, tdoc_type, tdoc_number, column_index, stats, failed_downloads, failed_extractions):
    target_release = str(row[6].value).strip() if row[6].value else ""
    spec_number = str(row[1].value).strip() if row[1].value else ""
    cr_number = str(row[2].value).strip() if row[2].value else ""
    title = str(row[7].value).strip() if row[7].value else ""
    tsg_src = str(row[15].value).strip() if row[15].value else ""

    formatted_name = f"{target_release}_{tsg_src}_TS-{spec_number}_CR-{cr_number}_{title}*_{tdoc_type}TDOC-{tdoc_number}"

    html_page_link = row[column_index].hyperlink.target if row[column_index].hyperlink else ""

    if html_page_link:
        download_link = extract_download_link(html_page_link)
        
        if download_link:
            filename = os.path.join('downloads', os.path.basename(download_link))
            os.makedirs('downloads', exist_ok=True)

            if os.path.exists(filename):
                logger.info(f"File already exists: {filename}")
                stats['existing'] += 1
            else:
                if download_file(download_link, filename):
                    logger.info(f"Downloaded: {filename}")
                    stats['downloaded'] += 1
                else:
                    logger.error(f"Failed to download: {download_link}")
                    failed_downloads.append(download_link)
                    stats['failed_downloads'] += 1
                    return

            if filename.lower().endswith('.zip'):
                extract_path = os.path.join('extracted', formatted_name)
                os.makedirs(extract_path, exist_ok=True)
                try:
                    extract_zip(filename, extract_path)
                    logger.info(f"Extracted to: {extract_path}")
                    stats['extracted'] += 1
                except Exception as e:
                    logger.error(f"Failed to extract: {filename}. Error: {str(e)}")
                    failed_extractions.append(filename)
            else:
                new_filename = os.path.join('extracted', formatted_name, os.path.basename(filename))
                os.makedirs(os.path.dirname(new_filename), exist_ok=True)
                shutil.copy2(filename, new_filename)
                logger.info(f"Copied to: {new_filename}")
                stats['copied'] += 1
        else:
            logger.warning(f"Skipping {tdoc_type} TDOC due to access issues or missing download link.")
            stats['skipped'] += 1

def main(excel_file, use_parallel=False):
    logger = setup_logging()
    
    stats = defaultdict(int)
    failed_downloads = []
    failed_extractions = []

    logger.info("Starting download process")

    workbook = openpyxl.load_workbook(excel_file)

    if use_parallel:
        with concurrent.futures.ThreadPoolExecutor(max_workers=5) as executor:
            futures = []
            for sheet in workbook.sheetnames:
                worksheet = workbook[sheet]
                for row in worksheet.iter_rows(min_row=2):
                    wg_tdoc = str(row[8].value).strip() if row[8].value else ""
                    tsg_tdoc = str(row[12].value).strip() if row[12].value else ""
                    
                    if wg_tdoc:
                        futures.append(executor.submit(process_tdoc, row, "WG", wg_tdoc, 8, stats, failed_downloads, failed_extractions))
                    if tsg_tdoc:
                        futures.append(executor.submit(process_tdoc, row, "TSG", tsg_tdoc, 12, stats, failed_downloads, failed_extractions))
            
            concurrent.futures.wait(futures)
    else:
        for sheet in workbook.sheetnames:
            worksheet = workbook[sheet]
            for row in worksheet.iter_rows(min_row=2):
                wg_tdoc = str(row[8].value).strip() if row[8].value else ""
                tsg_tdoc = str(row[12].value).strip() if row[12].value else ""
                
                if wg_tdoc:
                    process_tdoc(row, "WG", wg_tdoc, 8, stats, failed_downloads, failed_extractions)
                if tsg_tdoc:
                    process_tdoc(row, "TSG", tsg_tdoc, 12, stats, failed_downloads, failed_extractions)

    workbook.close()

    # Log statistics
    logger.info("\nDownload Statistics:")
    logger.info(f"Total files processed: {sum(stats.values())}")
    logger.info(f"Files already existing: {stats['existing']}")
    logger.info(f"Files downloaded: {stats['downloaded']}")
    logger.info(f"Files extracted: {stats['extracted']}")
    logger.info(f"Files copied (non-zip): {stats['copied']}")
    logger.info(f"Failed downloads: {stats['failed_downloads']}")
    logger.info(f"Skipped items: {stats['skipped']}")

    if failed_downloads:
        logger.error("\nFailed Downloads:")
        for url in failed_downloads:
            logger.error(url)

    if failed_extractions:
        logger.error("\nFailed Extractions:")
        for file in failed_extractions:
            logger.error(file)

    logger.info("Download process completed")

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Download and process 3GPP CRs")
    parser.add_argument('excel_file', nargs='?', default='CRs.xlsx', help='Path to the Excel file (default: CRs.xlsx)')
    parser.add_argument('--parallel', action='store_true', help='Enable parallel processing')
    args = parser.parse_args()

    try:
        main(args.excel_file, use_parallel=args.parallel)
    except KeyboardInterrupt:
        print("\nProgram interrupted by user. Exiting...")
